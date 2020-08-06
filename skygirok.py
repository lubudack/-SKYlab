import discord
import asyncio
import openpyxl
import datetime

client = discord.Client()

owner = [653075791814590487]

@client.event
async def on_message(message):
    if message.content.startswith('/잊어'):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 201):
            if sheet["A" + str(i)].value == str(memory[1]):
                sheet["A" + str(i)].value = "-"
                sheet["B" + str(i)].value = " "
                await message.channel.send("잊어버렸어요!")
                file.save('기억.xlsx')
                break

    if message.content.startswith("/배워"):
        file = openpyxl.load_workbook('기억.xlsx')
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 201):
            if sheet["A"+str(i)].value == "-":
                sheet["A" + str(i)].value = learn[1]
                sheet["B" + str(i)].value = learn[2]
                await message.channel.send("정상적으로 배웠어요!")
                await message.channel.send("★ 현재 사용중인 데이터 저장용량 : " + str(i)+" / 200 ★")
                break
        file.save("기억.xlsx")


    if message.content.startswith("카이야"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 201):
            if sheet["A" + str(i)].value == memory[1]:
                await message.channel.send(sheet["B" + str(i)].value)
                break

    if message.content.startswith("/기억 초기화") or message.content.startswith("/기억초기화"):
        if message.author.id in owner:
            file = openpyxl.load_workbook("기억.xlsx")
            sheet = file.active
            for i in range(1, 201):
                sheet["A"+str(i)].value = "-"
            embed = discord.Embed(title="SkyBOT : 기억초기화", description=f"Bot Developer 권한으로 기억을 바꿨어요!\n \n 기억초기화가 정상적으로 완료되었어요!", timestamp=message.created_at,
            colour = discord.Colour.teal()    
        )
            await message.channel.send(embed=embed)
            file.save("기억.xlsx")
        else:
            embed = discord.Embed(title=f"SkyBOT : Error", description=f"Bot Developer 등급보다 낮은 등급을 가지고 있습니다. \n \n {message.author.mention} 님의 등급 : User", timestamp=message.created_at,
            colour = discord.Colour.red()
    )
            embed.set_footer(text="Bot Developer 등급이상만 사용할 수 있는 명령입니다.")        
            await message.channel.send(embed=embed)

client.run('NzM2MTM1NjcyODQyNDIwMzI2.XxqZxQ.qxhev3KcOcRWwQvqZNs1n2mzkTI')